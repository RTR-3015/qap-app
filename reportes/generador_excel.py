"""
Generación de reportes en formato Excel para los resultados del QAP.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generar_reporte_excel(resultado):
    """
    Genera un archivo Excel en memoria (BytesIO) con el resumen de resultados,
    las matrices de entrada y las asignaciones inicial y final.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen QAP"

    titulo_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    ws["A1"] = "Reporte de Resultados - Problema de Asignación Cuadrática (QAP)"
    ws["A1"].font = titulo_font
    ws.merge_cells("A1:D1")

    filas_resumen = [
        ("Número de departamentos/ubicaciones (n)", resultado["n"]),
        ("", ""),
        ("Fase 1 - Construcción", ""),
        ("Método de construcción", resultado["metodo_construccion"]),
        ("Costo inicial (Z)", resultado["costo_inicial"]),
        ("Tiempo de ejecución (s)", resultado["tiempo_construccion"]),
        ("", ""),
        ("Fase 2 - Mejora", ""),
        ("Método de mejora", resultado["metodo_mejora"]),
        ("Costo final (Z)", resultado["costo_final"]),
        ("Tiempo de ejecución (s)", resultado["tiempo_mejora"]),
        ("Porcentaje de mejora (%)", resultado["porcentaje_mejora"]),
        ("Tiempo total de ejecución (s)", resultado["tiempo_total"]),
    ]

    fila_actual = 3
    for etiqueta, valor in filas_resumen:
        ws.cell(row=fila_actual, column=1, value=etiqueta).font = Font(bold=etiqueta in (
            "Fase 1 - Construcción", "Fase 2 - Mejora"
        ))
        ws.cell(row=fila_actual, column=2, value=valor)
        fila_actual += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30

    # Hoja de asignaciones
    ws2 = wb.create_sheet("Asignaciones")
    n = resultado["n"]

    ws2.cell(row=1, column=1, value="Departamento").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Ubicación (Inicial)").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Ubicación (Final)").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill

    for i in range(n):
        ws2.cell(row=i + 2, column=1, value=f"D{i+1}")
        ws2.cell(row=i + 2, column=2, value=f"U{resultado['asignacion_inicial'][i] + 1}")
        ws2.cell(row=i + 2, column=3, value=f"U{resultado['asignacion_final'][i] + 1}")

    for col in ("A", "B", "C"):
        ws2.column_dimensions[col].width = 20

    # Hoja de matrices de entrada
    ws3 = wb.create_sheet("Matrices de entrada")
    _escribir_matriz(ws3, "Matriz de Flujo", resultado["flujo"], fila_inicio=1)
    _escribir_matriz(ws3, "Matriz de Distancia", resultado["distancia"], fila_inicio=n + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _escribir_matriz(ws, titulo, matriz, fila_inicio):
    ws.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True)
    n = len(matriz)
    for j in range(n):
        ws.cell(row=fila_inicio + 1, column=j + 2, value=f"U{j+1}").font = Font(bold=True)
    for i in range(n):
        ws.cell(row=fila_inicio + 2 + i, column=1, value=f"D{i+1}").font = Font(bold=True)
        for j in range(n):
            ws.cell(row=fila_inicio + 2 + i, column=j + 2, value=matriz[i][j])
